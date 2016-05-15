################################################################################
# Global "import" modules, enabling access to Postgres function library        #
################################################################################

# Command line argument handling functions
import sys
import os, os.path
import csv

import openpyxl as op



wb = op.Workbook()

c = open('test2.csv','rb')   
reader = csv.reader(c)
rows = list(reader)
ws = wb.get_sheet_by_name('Sheet')


for j in range(0,len(rows)):
    for k in range(0,len(rows[j])):
        try:
            ws.cell(row=j+1, column=k+1).value = float(rows[j][k]) 
        except:
            ws.cell(row=j+1, column=k+1).value = rows[j][k] 
        #ws.cell(row=j+1, column=k+1).value = rows[j][k] 
        
        
wb.save('testoutput2.xlsx')

# TEST CHANGING FILE
